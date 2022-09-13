import time
import os
import openpyxl
import datetime

from datetime import datetime, timedelta
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, NamedStyle
                        )
from copy import copy
from pathlib import Path

number_row_1 = 0
date_old = datetime
date_old_min = datetime
name_sheet = ''
ws_1 = ''
ws_2 = ''
wb_1 = ''
wb_2 = ''
ws_1_temp = ''
row_min_del = 1
number_row = 12
row_limit = 0
path_1 = None
path_2 = None
path_file_1 = None
path_file_2 = None
m = 0
temp = ''
date_check = ''
date_start_input = ''
date_end_input = ''
min_row = 0
max_row = 0
flag_finish = 0
d1 = None
d2 = None
d3 = None
r = 1

path_names_workers = Path('Names.txt')
names_workers = path_names_workers.read_text().splitlines()

names_clear = str(path_names_workers.read_text().splitlines()).replace('[', '\n ').replace(']', '').replace("'", '').replace(',', '\n')

path_names_ceo = Path('Names.txt')
names_ceo = path_names_ceo.read_text().splitlines()

names_clear_ceo = str(path_names_ceo.read_text().splitlines()).replace('[', '\n ').replace(']', '').replace("'", '').replace(',', '\n')


list_num = list(range(1, 501))

Col_A = NamedStyle(name="Col_A")
Col_A.font = Font(name='Calibri',
                 size=11,
                 bold=True)
Col_A.alignment = Alignment(horizontal='center')
Col_A.fill = PatternFill(fill_type='solid', fgColor='FF0000')
Col_A.border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )


Col_A_pol_font = Font(name='Calibri',
                 size=11)
Col_A_pol_alignment = Alignment(horizontal='center')
Col_A_pol_border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )
Col_A_fill = PatternFill(
            fill_type='solid', fgColor='FF0000'
            )
Col_A_font = Font(name='Calibri',
                 size=11,
                 bold=True)

Col_B_font = Font(name='Calibri',
                 size=11,
                 bold=True,
                 color="FF0000")
Col_B_alignment = Alignment(horizontal='center')
Col_B_border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )


Col_C_font = Font(name='Calibri',
                 size=11,
                 bold=True)
Col_C_alignment = Alignment(horizontal='center')
Col_C_border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )


Col_D_font = Font(name='Calibri',
                 size=11,
                 bold=True)
Col_D_fill = PatternFill(fill_type='solid', fgColor='e98244')
Col_D_alignment = Alignment(horizontal='center')
Col_D_border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )


Col_E_minus_font = Font(name='Calibri',
                 size=11,
                 bold=True,
                 color="4eac5b")
Col_E_minus_alignment = Alignment(horizontal='center')
Col_E_minus_border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )


Col_E_plus_font = Font(name='Calibri',
                 size=11,
                 bold=True,
                 color="FF0000")
Col_E_plus_alignment = Alignment(horizontal='center')
Col_E_plus_border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'),
            )


def excel_create(path_file_1, path_file_2):
    global wb_1, wb_2, ws_1, ws_2, path_1, path_2

    path_1 = fr'{path_file_1}'
    path_2 = fr'{path_file_2}'

    wb_1 = openpyxl.load_workbook(filename=path_1, read_only=False)
    wb_2 = openpyxl.load_workbook(filename=path_2, read_only=False)

    ws_1 = wb_1['Sheet0']
    ws_2 = wb_2.active

    today_sheet()


def score():

    number_1 = 0
    number_2 = 0
    number_3 = 0
    number_4 = 0
    number_5 = 0
    number_row = 0

    for row in ws_1.rows:
        number_row += 1
        if ws_1[f'R{number_row}'].value == '1' and ws_1[f'W{number_row}'].value in names_workers:
            number_1 += 1

        if ws_1[f'R{number_row}'].value == '2' and ws_1[f'W{number_row}'].value in names_workers:
            number_2 += 1

        if ws_1[f'R{number_row}'].value == '3' and ws_1[f'W{number_row}'].value in names_workers:
            number_3 += 1

        if ws_1[f'R{number_row}'].value == '4' and ws_1[f'W{number_row}'].value in names_workers:
            number_4 += 1

        if ws_1[f'R{number_row}'].value == '5' and ws_1[f'W{number_row}'].value in names_workers:
            number_5 += 1


    num_1_old = ws_2['S3'].value
    num_2_old = ws_2['Q3'].value
    num_3_old = ws_2['O3'].value
    num_4_old = ws_2['M3'].value
    num_5_old = ws_2['K3'].value

    if num_1_old == number_1:
        ws_2['T3'] = ''
    else:
        if int(num_1_old) > number_1:
            text_1 = int(num_1_old) - int(number_1)
            ws_2['T3'] = f'-{text_1}'
            ws_2['S3'] = number_1
            ws_2['T3'].font = Font(color="4eac5b")
        else:
            text_1 = int(number_1) - int(num_1_old)
            ws_2['T3'] = f'+{text_1}'
            ws_2['S3'] = number_1
            ws_2['T3'].font = Font(color="FF0000")

    if num_2_old == number_2:
        ws_2['R3'] = ''
    else:
        if int(num_2_old) > number_2:
            text_2 = int(num_2_old) - int(number_2)
            ws_2['R3'] = f'-{text_2}'
            ws_2['Q3'] = number_2
            ws_2['R3'].font = Font(color="4eac5b")
        else:
            text_2 = int(number_2) - int(num_2_old)
            ws_2['R3'] = f'+{text_2}'
            ws_2['Q3'] = number_2
            ws_2['R3'].font = Font(color="FF0000")

    if num_3_old == number_3:
        ws_2['P3'] = ''
    else:
        if int(num_3_old) > number_3:
            text_3 = int(num_3_old) - int(number_3)
            ws_2['P3'] = f'-{text_3}'
            ws_2['O3'] = number_3
            ws_2['P3'].font = Font(color="4eac5b")
        else:
            text_3 = int(number_3) - int(num_3_old)
            ws_2['P3'] = f'+{text_3}'
            ws_2['O3'] = number_3
            ws_2['P3'].font = Font(color="FF0000")

    if num_4_old == number_4:
        ws_2['N3'] = ''
    else:
        if int(num_4_old) > number_4:
            text_4 = int(num_4_old) - int(number_4)
            ws_2['N3'] = f'-{text_4}'
            ws_2['M3'] = number_4
            ws_2['N3'].font = Font(color="4eac5b")
        else:
            text_4 = int(number_4) - int(num_4_old)
            ws_2['N3'] = f'+{text_4}'
            ws_2['M3'] = number_4
            ws_2['N3'].font = Font(color="4eac5b")

    if num_5_old == number_5:
        ws_2['L3'] = ''
    else:
        if int(num_5_old) > number_5:
            text_5 = int(num_5_old) - int(number_5)
            ws_2['L3'] = f'-{text_5}'
            ws_2['K3'] = number_5
            ws_2['L3'].font = Font(color="FF0000")
        else:
            text_5 = int(number_5) - int(num_5_old)
            ws_2['L3'] = f'+{text_5}'
            ws_2['K3'] = number_5
            ws_2['L3'].font = Font(color="4eac5b")

    return 'Оценки успешно записаны!'


def all_values():

    number_row = 0
    values = 0

    for row in ws_1.rows:
        number_row += 1
        if ws_1[f'Q{number_row}'].value == 'Решено' and ws_1[f'W{number_row}'].value in names_workers:
            values += 1

    text_all = values - 3

    ws_2['I3'] = text_all

    wb_1.save(path_1)
    wb_2.save(path_2)


def today_sheet():
    global ws_2, wb_2

    name_sheet = str(ws_2).replace('<', '').replace('>', '').replace('"', '').replace('Worksheet', '').lstrip()
    today = str(datetime.today().strftime("Мониторинг от %d.%m.%Y|%H.%M"))

    print(today)

    if name_sheet != today:
        ws_2_old = ws_2
        wb_2.create_sheet(index=1, title=today)
        ws_2 = wb_2[today]
        wb_2.active = wb_2[today]
        new_sheet(ws_2_old)


def new_sheet(ws_2_old):
    global ws_2

    number_row = 0

    width_row_1 = 11
    width_row = 4.5

    ws_2.column_dimensions['A'].width = width_row_1
    ws_2.column_dimensions['B'].width = width_row_1
    ws_2.column_dimensions['C'].width = width_row_1
    ws_2.column_dimensions['D'].width = width_row_1
    ws_2.column_dimensions['E'].width = width_row_1
    ws_2.column_dimensions['H'].width = 14
    ws_2.column_dimensions['K'].width = width_row
    ws_2.column_dimensions['L'].width = width_row
    ws_2.column_dimensions['M'].width = width_row
    ws_2.column_dimensions['N'].width = width_row
    ws_2.column_dimensions['O'].width = width_row
    ws_2.column_dimensions['P'].width = width_row
    ws_2.column_dimensions['Q'].width = width_row
    ws_2.column_dimensions['R'].width = width_row
    ws_2.column_dimensions['S'].width = width_row
    ws_2.column_dimensions['T'].width = width_row

    for row in ws_2_old.rows:
        number_row += 1
        for cell in row:
            col_number = cell.column
            cell_info = cell.value

            c1 = ws_2.cell(row=number_row, column=col_number)
            c1.value = cell_info

            if cell.has_style:
                c1._style = copy(cell._style)

    ws_2.merge_cells('A11:E11')
    ws_2.merge_cells('K2:L2')
    ws_2.merge_cells('M2:N2')
    ws_2.merge_cells('O2:P2')
    ws_2.merge_cells('Q2:R2')
    ws_2.merge_cells('S2:T2')


def insert_workers():
    global ws_1_temp
    number_row = 0
    i = 1

    wb_1.create_sheet(index=1, title='workers_1')
    ws_1_temp = wb_1['workers_1']

    for row in ws_1.rows:
        number_row += 1
        if ws_1[f'W{number_row}'].value in names_workers and ws_1[f'Y{number_row}'].value in names_ceo or ws_1[f'W{number_row}'].value in names_workers or ws_1[f'Y{number_row}'].value in names_ceo:
            worker_1 = ws_1[f'J{number_row}'].value
            ws_1_temp.cell(row=i, column=1).value = worker_1
            i += 1
    date_count()
    wb_1.save(path_1)

# def insert_ceo():


def date_first():
    global ws_1_temp
    number_row = 0
    i = 1

    wb_1.create_sheet(index=1, title='test_1')
    ws_1_temp = wb_1['test_1']

    for row in ws_1.rows:
        number_row += 1
        if ws_1[f'N{number_row}'].value != 'Отправлен ответ заявителю' and ws_1[f'S{number_row}'].value != '-' and ws_1[f'W{number_row}'].value in names_workers:
            date_entry_1 = ws_1[f'J{number_row}'].value
            date_temp = datetime.strptime(date_entry_1, '%d.%m.%Y')
            ws_1_temp.cell(row=i, column=1).value = date_temp
            i += 1
    date_count()
    wb_1.save(path_1)


def date_first_1():
    number_row_1 = 1

    for row in ws_1_temp.rows:
        date_entry_1 = ws_1_temp[f'A{number_row_1}'].value
        count_date = ws_1_temp[f'B{number_row_1}'].value

        number_row_1 += 1
        if date_entry_1 is not None:
            date_second(date_entry_1, count_date)


def date_delete_empty_rows():
    global row_limit

    n = 13

    while n <= row_limit:
        if ws_2.cell(row=n, column=6).value != 'new':
            if ws_2.cell(row=n, column=3).value is None:
                if ws_2.cell(row=n, column=4).value is not None:
                    temp_cell = ws_2.cell(row=n, column=4).value
                    ws_2.cell(row=n, column=4).value = ''
                    ws_2.cell(row=n, column=5).value = f'-{temp_cell}'
                    n += 1
                else:
                    ws_2.delete_rows(n, 1)
                    row_limit = row_limit - 1
            else: #ЧТО-ТО ТУТ НЕ ПУСКАЕТ ДАЛЬШЕ
                if ws_2.cell(row=n, column=4).value is not None:
                    temp_cell = ws_2.cell(row=n, column=4).value
                    ws_2.cell(row=n, column=4).value = ''

                    if ws_2.cell(row=n, column=5).value is not None:
                        if ws_2[f'E{number_row}'].value.find("/") != -1:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[2]

                            if text.find("-") != -1:
                                text_new = f'{temp_cell}' + f'{text}'
                                ws_2[f'E{number_row}'].value = f'-{text_new}'
                                n += 1

                            else:
                                text_new = f'-{temp_cell}/{text}'
                                ws_2[f'E{number_row}'].value = text_new
                                n += 1
                        else:
                            ws_2[f'E{number_row}'].value = f'-{temp_cell}'
                            n += 1
                    else:
                        ws_2[f'E{number_row}'].value = f'-{temp_cell}'
                        n += 1
                else:
                    n += 1
                    #СТРАННАЯ ШТУКА С УДАЛЕНИЕМ И УЗНАТЬ АКТУАЛЬНЫЙ СПИСОК ФИО
        else:
            n += 1


def date_second(date_entry_1, count_date):
    global row_limit
    number_row = 13

    date_temp = datetime.strptime('04.05.2000', '%d.%m.%Y')
    date_temp_1 = datetime.strptime('04.05.2022', '%d.%m.%Y')
    diap_2 = abs(date_temp - date_temp_1)

    while number_row <= ws_2.max_row:
        date_11 = ws_2[f'A{number_row}'].value

        if date_11 is not None:
            try:
                diap_1 = abs(date_11 - date_entry_1)
            except:
                date_11 = datetime.strptime(date_11, '%d.%m.%Y')
                diap_1 = abs(date_11 - date_entry_1)

            if diap_1 < diap_2:
                diap_2 = diap_1
                diap_row = number_row

            if date_11 == date_entry_1:
                if ws_2[f'D{number_row}'].value == count_date:
                    if ws_2[f'E{number_row}'].value is not None:
                        if ws_2[f'E{number_row}'].value.find("/") != -1:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[2]
                            ws_2[f'E{number_row}'].value = text
                        else:
                            ws_2[f'E{number_row}'].value = None
                    ws_2[f'F{number_row}'].value = 'new'

                    row_limit = number_row

                    return

                elif ws_2[f'D{number_row}'].value is None:
                    ws_2[f'D{number_row}'].value = count_date

                    if ws_2[f'E{number_row}'].value is not None:

                        if ws_2[f'E{number_row}'].value.find("/") != -1:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[2]

                            if temp_text.find("+") != -1:
                                text_new = f'{count_date}' + f'{text}'
                                ws_2[f'E{number_row}'].value = f'+{text_new}'

                            else:
                                text_new = f'+{count_date}/{text}'
                                ws_2[f'E{number_row}'].value = text_new
                        else:
                            ws_2[f'E{number_row}'].value = f'+{count_date}'
                    else:
                        ws_2[f'E{number_row}'].value = f'+{count_date}'

                    ws_2[f'F{number_row}'].value = 'new'

                    row_limit = number_row

                    return

                elif ws_2[f'D{number_row}'].value > count_date:
                    count_date_old = ws_2[f'D{number_row}'].value
                    minus = int(count_date_old) - count_date
                    ws_2[f'D{number_row}'].value = count_date

                    if ws_2[f'E{number_row}'].value is not None:

                        if ws_2[f'E{number_row}'].value.find("/") != -1:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[2]

                            if text.find("-") != -1:
                                text_new = f'-{minus}' + f'{text}'
                                ws_2[f'E{number_row}'].value = f'-{text_new}'

                            else:
                                text_new = f'-{minus}/{text}'
                                ws_2[f'E{number_row}'].value = text_new
                        else:
                            ws_2[f'E{number_row}'].value = f'-{minus}'
                    else:
                        ws_2[f'E{number_row}'].value = f'-{minus}'
                    ws_2[f'F{number_row}'].value = 'new'

                    row_limit = number_row

                    return

                elif ws_2[f'D{number_row}'].value < count_date:
                    count_date_old = ws_2[f'D{number_row}'].value
                    plus = count_date - int(count_date_old)
                    ws_2[f'D{number_row}'].value = count_date

                    if ws_2[f'E{number_row}'].value is not None:

                        if ws_2[f'E{number_row}'].value.find("/") != -1:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[2]

                            if text.find("+") != -1:
                                text_new = f'{plus}' + f'+{text}'
                                ws_2[f'E{number_row}'].value = f'+{text_new}'

                            else:
                                text_new = f'+{plus}/{text}'
                                ws_2[f'E{number_row}'].value = text_new
                        else:
                            ws_2[f'E{number_row}'].value = f'+{plus}'
                    else:
                        ws_2[f'E{number_row}'].value = f'+{plus}'
                    ws_2[f'F{number_row}'].value = 'new'

                    row_limit = number_row

                    return
            else:
                number_row += 1
        else:
            break

    ws_2.insert_rows(diap_row, 1)

    date_new_1 = date_entry_1 + timedelta(days=5)
    date_new = date_new_1.strftime("%d.%m.%Y")
    date_str = date_entry_1.strftime("%d.%m.%Y")

    ws_2[f'A{diap_row}'] = date_str
    ws_2[f'B{diap_row}'] = date_new
    ws_2[f'D{diap_row}'] = count_date
    ws_2[f'E{diap_row}'] = f'+{count_date}'
    ws_2[f'F{diap_row}'].value = 'new'

    row_limit = number_row

    ws_2[f'A{diap_row}'].fill = Col_A_fill
    ws_2[f'A{diap_row}'].font = Col_A_font
    ws_2[f'A{diap_row}'].alignment = Col_A_pol_alignment
    ws_2[f'A{diap_row}'].border = Col_A_pol_border
    ws_2[f'B{diap_row}'].font = Col_B_font
    ws_2[f'B{diap_row}'].alignment = Col_B_alignment
    ws_2[f'B{diap_row}'].border = Col_B_border
    ws_2[f'D{diap_row}'].font = Col_D_font
    ws_2[f'D{diap_row}'].alignment = Col_D_alignment
    ws_2[f'D{diap_row}'].border = Col_D_border
    ws_2[f'D{diap_row}'].fill = Col_D_fill
    ws_2[f'E{diap_row}'].font = Col_E_plus_font
    ws_2[f'E{diap_row}'].alignment = Col_E_plus_alignment
    ws_2[f'E{diap_row}'].border = Col_E_plus_border


def formula():
    row_max = ws_2.max_row
    row_minus = row_max-1

    cell_C = ws_2.cell(row=row_max, column=3)
    cell_D = ws_2.cell(row=row_max, column=4)
    cell_E = ws_2.cell(row=row_max, column=5)

    cell_C.value = f"=SUM(C13:C{row_minus})"
    cell_D.value = f"=SUM(D13:D{row_minus})"
    cell_E.value = f"=SUM(C{row_max}:D{row_max})"


def date_delete_all():
    i = 1

    while i <= ws_1_temp.max_row:
        if ws_1_temp.cell(row=i, column=1).value != None:
            ws_1_temp.delete_rows(i, 1)
        else:
            i += 1
    wb_1.save(path_1)


def copy_second_date():
    number_row = 0
    i = 1

    for row in ws_1.rows:
        number_row += 1
        if ws_1.cell(row=number_row, column=17).value !='Решено' and ws_1.cell(row=number_row, column=19).value == '-' and ws_1.cell(row=number_row, column=23).value in names_workers:
            date_entry_1 = ws_1.cell(row=number_row, column=10).value
            date_temp = datetime.strptime(date_entry_1, '%d.%m.%Y')
            ws_1_temp.cell(row=i, column=1).value = date_temp
            i += 1
    date_count()


def date_count():
    count = 0
    row_min = 0
    n = 1

    for row in ws_1_temp['A']:
        row_min += 1
        temp = ws_1_temp[f'A{row_min}'].value

        if temp is not None:
            while n <= ws_1_temp.max_row:
                value_check = ws_1_temp[f'A{n}'].value
                if temp == value_check:
                    count += 1
                    n += 1
                else:
                    n += 1
            n = 1
            ws_1_temp[f'B{row_min}'].value = count
            count = 0


def delete():
    global list_num
    i = 1
    row_min = 0
    count = 0

    for row in ws_1_temp['A']:
        row_min += 1
        temp = ws_1_temp[f'A{row_min}'].value

        if ws_1_temp[f'B{row_min}'].value is not None:
            count_main = ws_1_temp[f'B{row_min}'].value

            while i <= ws_1_temp.max_row:
                value_check = ws_1_temp[f'A{i}'].value
                if value_check == temp and count < count_main and row_min != i:
                    ws_1_temp.delete_rows(i, 1)
                    count += 1
                else:
                    i += 1
            count = 0
            i = 1


def second_eteration_date():
    number_row_1 = 1

    for row in ws_1_temp.rows:
        date_entry_1 = ws_1_temp[f'A{number_row_1}'].value
        count_date = ws_1_temp[f'B{number_row_1}'].value
        number_row_1 += 1

        if date_entry_1 is not None:
            date_str = date_entry_1.strftime("%d.%m.%Y")

            if date_str != '09.06.2022':
                date_second_eteration_paste(date_entry_1, count_date)


def date_second_eteration_paste(date_entry_1, count_date):
    global row_limit
    number_row = 13

    date_temp = datetime.strptime('04.05.2000', '%d.%m.%Y')
    date_temp_1 = datetime.strptime('04.05.2022', '%d.%m.%Y')
    diap_2 = abs(date_temp - date_temp_1)

    while number_row <= ws_2.max_row:
        date_11 = ws_2[f'A{number_row}'].value

        if date_11 is not None:
            try:
                diap_1 = abs(date_11 - date_entry_1)
            except:
                date_11 = datetime.strptime(date_11, '%d.%m.%Y')
                diap_1 = abs(date_11 - date_entry_1)

            if diap_1 < diap_2:
                diap_2 = diap_1
                diap_row = number_row + 1

            if date_11 == date_entry_1:
                if ws_2[f'C{number_row}'].value == count_date:
                    if ws_2[f'E{number_row}'].value is not None:
                        if ws_2[f'E{number_row}'].value.find("/") != -1:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[0]
                            ws_2[f'E{number_row}'].value = text
                        if ws_2[f'D{number_row}'].value is None:
                            ws_2[f'E{number_row}'].value = None
                    else:
                        ws_2[f'E{number_row}'].value = None

                    row_limit = number_row
                    ws_2[f'G{number_row}'].value = 'new_2'

                    return

                elif ws_2[f'C{number_row}'].value is None:
                    ws_2[f'C{number_row}'].value = count_date

                    if ws_2[f'E{number_row}'].value is None:
                        ws_2[f'E{number_row}'].value = f'+{count_date}'
                    else:
                        temp_text = ws_2[f'E{number_row}'].value

                        if temp_text.find("+") != -1:
                            text_new = temp_text + count_date
                            ws_2[f'E{number_row}'].value = text_new

                        else:
                            text_new = f'-{temp_text}/+{count_date}'
                            ws_2[f'E{number_row}'].value = text_new

                    row_limit = number_row
                    ws_2[f'G{number_row}'].value = 'new_2'

                    return

                elif ws_2[f'C{number_row}'].value > count_date:
                    count_date_old = ws_2[f'C{number_row}'].value
                    minus = int(count_date_old) - count_date

                    ws_2[f'C{number_row}'].value = count_date

                    if ws_2[f'F{number_row}'].value == 'new':
                        if ws_2[f'E{number_row}'].value is not None:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[0]
                            if text.find("-") != -1:
                                text_new = int(text)-int(minus)
                                ws_2[f'E{number_row}'].value = f'{text_new}'
                            else:
                                text_new = f'{text}/-{minus}'
                                ws_2[f'E{number_row}'].value = text_new
                        else:
                            ws_2[f'E{number_row}'].value = f'-{minus}'
                    else:
                        ws_2[f'E{number_row}'].value = f'-{minus}'

                    row_limit = number_row
                    ws_2[f'G{number_row}'].value = 'new_2'

                    return

                elif ws_2[f'C{number_row}'].value < count_date:
                    count_date_old = ws_2[f'C{number_row}'].value
                    plus = count_date - int(count_date_old)

                    ws_2[f'C{number_row}'].value = count_date

                    if ws_2[f'F{number_row}'].value == 'new':
                        if ws_2[f'E{number_row}'].value is not None:
                            temp_text = ws_2[f'E{number_row}'].value
                            text = temp_text.partition('/')[0]

                            if text.find("+") != -1:
                                text_new = int(text) + int(plus)
                                ws_2[f'E{number_row}'].value = f'+{text_new}'

                            else:
                                text_new = f'{text}/+{plus}'
                                ws_2[f'E{number_row}'].value = text_new
                        else:
                            ws_2[f'E{number_row}'].value = f'+{plus}'
                    else:
                        ws_2[f'E{number_row}'].value = f'+{plus}'

                    row_limit = number_row
                    ws_2[f'G{number_row}'].value = 'new_2'

                    return
            else:
                number_row += 1
        else:
            break

    ws_2.insert_rows(diap_row, 1)

    date_new_1 = date_entry_1 + timedelta(days=5)
    date_new = date_new_1.strftime("%d.%m.%Y")
    date_str = date_entry_1.strftime("%d.%m.%Y")

    ws_2[f'A{diap_row}'] = date_str
    ws_2[f'B{diap_row}'] = date_new
    ws_2[f'C{diap_row}'] = count_date
    ws_2[f'E{diap_row}'] = f'+{count_date}'

    row_limit = number_row

    ws_2[f'A{diap_row}'].font = Col_A_pol_font
    ws_2[f'A{diap_row}'].alignment = Col_A_pol_alignment
    ws_2[f'A{diap_row}'].border= Col_A_pol_border
    ws_2[f'B{diap_row}'].font = Col_B_font
    ws_2[f'B{diap_row}'].alignment = Col_B_alignment
    ws_2[f'B{diap_row}'].border = Col_B_border
    ws_2[f'C{diap_row}'].font = Col_C_font
    ws_2[f'C{diap_row}'].alignment = Col_C_alignment
    ws_2[f'C{diap_row}'].border = Col_C_border
    ws_2[f'D{diap_row}'].font = Col_D_font
    ws_2[f'D{diap_row}'].alignment = Col_D_alignment
    ws_2[f'D{diap_row}'].border = Col_D_border
    ws_2[f'D{diap_row}'].fill = Col_D_fill
    ws_2[f'E{diap_row}'].font = Col_E_plus_font
    ws_2[f'E{diap_row}'].alignment = Col_E_plus_alignment
    ws_2[f'E{diap_row}'].border = Col_E_plus_border

    ws_2[f'G{number_row}'].value = 'new_2'


def date_delete_empty_rows_second():
    global row_limit

    n = 13

    while n <= row_limit:
        ws_2.cell(row=n, column=6).value = None
        if ws_2.cell(row=n, column=7).value != 'new_2' and ws_2.cell(row=n, column=3).value is not None:
            minus = ws_2.cell(row=n, column=3).value
            ws_2.cell(row=n, column=3).value = None

            if ws_2[f'E{n}'].value is not None:
                temp_text = ws_2[f'E{n}'].value
                text = temp_text.partition('/')[0]
                if text.find("-") != -1:
                    text_new = int(text) - int(minus)
                    ws_2[f'E{n}'].value = f'{text_new}'
                else:
                    text_new = f'{text}/-{minus}'
                    ws_2[f'E{n}'].value = text_new
            else:
                ws_2[f'E{n}'].value = f'-{minus}'
            n += 1
        else:
            ws_2.cell(row=n, column=7).value = None
            n += 1


def plus_minus_color():
    number_row = 13

    while number_row <= ws_2.max_row:
        if ws_2[f'E{number_row}'].value is not None:
            if ws_2[f'E{number_row}'].value.find("-") != -1:
                ws_2[f'E{number_row}'].font = Font(bold=True, color='4eac5b')
                number_row += 1
            elif ws_2[f'E{number_row}'].value.find("+") != -1:
                ws_2[f'E{number_row}'].font = Font(bold=True, color='FF0000')
                number_row += 1
            else:
                number_row += 1
        else:
            number_row += 1


def main(file_1, file_2):
    global path_file_1, path_file_2
    path_file_1 = file_1
    path_file_2 = file_2
    excel_create(path_file_1, path_file_2)
    print('Программа начинает работу...')
    score()
    print('Значения')
    all_values()
    print('Сумма')
    date_first()
    delete()
    date_first_1()
    date_delete_empty_rows()
    date_delete_all()
    copy_second_date()
    delete()
    second_eteration_date()
    date_delete_empty_rows_second()
    formula()
    print('Формула')
    plus_minus_color()
    print('Цвета')

    del wb_1['test_1']

    wb_1.save(path_1)
    wb_2.save(path_2)
